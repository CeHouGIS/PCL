#!/usr/bin/env python3
"""Rebuild the 138-city workbook as five non-exclusive interaction layers.

The source evidence sheets are retained in the output workbook so every binary
edge remains traceable.  A city pair may occur in more than one layer.
"""

from __future__ import annotations

from collections import defaultdict
import os
from pathlib import Path
import shutil
import tempfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "data" / "raw" / "historical_city_connection_layers_138x138.xlsx"

CATEGORIES = {
    "C1": "Colonial and imperial cultural transmission",
    "C2": "Religious and civilizational diffusion",
    "C3": "Planning and architectural diffusion",
    "C4": "Shared official language",
    "C5": "Trade-mediated cultural exchange",
}

MATRIX_SHEETS = {
    "C1": "Colonial Imperial Matrix",
    "C2": "Religious Civiliz Matrix",
    "C3": "Planning Architecture Matrix",
    "C4": "Official Language Matrix",
    "C5": "Trade Exchange Matrix",
}

# These rules classify existing, sourced records.  They do not infer new ties
# from geographic proximity or from the learned image representations.
DIRECT_CODE_TO_CATEGORY = {
    "D1": "C1",  # colonial administrative transfer
    "D2": "C3",  # explicit planning transfer
    "D3": "C3",  # planner/engineer mobility
    "D4": "C1",  # imperial legal/institutional transfer in this evidence set
    "D5": "C2",  # migration-mediated cultural/civilizational diffusion
    "D6": "C3",  # occupation/reconstruction planning transfer
    "D7": "C3",  # construction/architectural transfer
    "D8": "C3",  # infrastructure model transfer
    "D9": "C5",  # port/municipal exchange in this evidence set
    "D10": "C2",  # direct religious or cultural-spatial transfer
}

COLONIAL_IMPERIAL_REGIMES = {
    "Portuguese Empire",
    "French Colonial Empire",
    "Dutch Empire / VOC",
    "German Colonial Empire",
    "Spanish Empire",
    "British Empire",
    "Belgian Colonial Regime",
    "Japanese Imperial / Occupation Regime",
    "United States Colonial / Occupation Regime",
    "Ottoman Imperial Regime",
}

# These memberships operationalize broad, historically shared civilizational
# or ideological spheres. Ottoman pairs can be multi-labelled C1 and C2.
CIVILIZATIONAL_REGIMES = {
    "Ottoman Imperial Regime",
    "Soviet / Eastern Bloc Planning Regime",
}

# VOC is explicitly a chartered trading-company network; D9 supplies directly
# documented port/municipal exchange.
TRADE_REGIMES = {"Dutch Empire / VOC"}


def sheet_rows(ws):
    """Return a worksheet as a list of dictionaries."""
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []
    headers = list(values[0])
    return [dict(zip(headers, row)) for row in values[1:] if any(v is not None for v in row)]


def add_edge(store, category, city_a, city_b, basis, detail, source_url=None, confidence=None):
    if not city_a or not city_b or city_a == city_b:
        return
    pair = tuple(sorted((str(city_a), str(city_b))))
    record = store[(category, pair)]
    record["bases"].add(str(basis))
    if detail:
        record["details"].add(str(detail))
    if source_url:
        record["urls"].add(str(source_url))
    if confidence:
        record["confidences"].add(str(confidence))


def copy_table(source_ws, target_ws):
    for row in source_ws.iter_rows(values_only=True):
        target_ws.append(list(row))


def style_table(ws, freeze="A2", filter_table=True):
    ws.freeze_panes = freeze
    if filter_table and ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="243447")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


def style_matrix(ws):
    style_table(ws, freeze="B2", filter_table=False)
    ws.column_dimensions["A"].width = 22
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 4.2
        ws.cell(1, col).alignment = Alignment(text_rotation=90, horizontal="center", vertical="bottom")
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).font = Font(bold=True)


def add_matrix_sheet(workbook, title, cities, edges, diagonal=0):
    ws = workbook.create_sheet(title)
    ws.append(["city", *cities])
    for city_a in cities:
        row = [city_a]
        for city_b in cities:
            row.append(diagonal if city_a == city_b else int(tuple(sorted((city_a, city_b))) in edges))
        ws.append(row)
    style_matrix(ws)


def main():
    source = load_workbook(WORKBOOK, data_only=True)
    required = {
        "City List", "Official Languages", "Language Tie Pairs",
        "Regime Memberships", "Shared Regime Pairs",
    }
    missing = required.difference(source.sheetnames)
    if missing:
        raise RuntimeError(f"Workbook is missing required source sheets: {sorted(missing)}")

    city_rows = sheet_rows(source["City List"])
    cities = [row["city"] for row in city_rows]
    city_set = set(cities)
    edges = defaultdict(lambda: {
        "bases": set(), "details": set(), "urls": set(), "confidences": set()
    })

    # Classify the two original direct-evidence tables. On later runs, read the
    # already harmonized Evidence Catalog, which makes this rebuild idempotent.
    direct_records = []
    seen_direct = set()
    if "Direct Evidence" in source.sheetnames and "Additional Direct Evidence" in source.sheetnames:
        direct_sources = [
            (sheet_name, sheet_rows(source[sheet_name]))
            for sheet_name in ("Direct Evidence", "Additional Direct Evidence")
        ]
    elif "Evidence Catalog" in source.sheetnames:
        reconstructed = []
        for catalog_row in sheet_rows(source["Evidence Catalog"]):
            reconstructed.append({
                "source_sheet": catalog_row.get("source_sheet") or "Evidence Catalog",
                "source_city": catalog_row.get("source_city"),
                "target_city": catalog_row.get("target_city"),
                "tie_code": catalog_row.get("original_tie_code"),
                "evidence_summary": catalog_row.get("evidence_summary"),
                "confidence": catalog_row.get("confidence"),
                "source_url": catalog_row.get("source_url"),
            })
        direct_sources = [("Evidence Catalog", reconstructed)]
    else:
        raise RuntimeError("Workbook has neither the original direct-evidence sheets nor Evidence Catalog")

    legacy_direct_pairs = set()
    for container_name, records in direct_sources:
        for row in records:
            sheet_name = row.get("source_sheet") or container_name
            category = DIRECT_CODE_TO_CATEGORY.get(row.get("tie_code"))
            if not category:
                continue
            city_a, city_b = row.get("source_city"), row.get("target_city")
            if city_a not in city_set or city_b not in city_set:
                raise RuntimeError(f"Unknown city in {sheet_name}: {city_a!r}, {city_b!r}")
            add_edge(
                edges, category, city_a, city_b,
                f"{sheet_name}: {row.get('tie_code')}", row.get("evidence_summary"),
                row.get("source_url"), row.get("confidence"),
            )
            legacy_direct_pairs.add(tuple(sorted((city_a, city_b))))
            key = (
                city_a, city_b, row.get("tie_code"), row.get("evidence_summary"),
                row.get("source_url"),
            )
            if key not in seen_direct:
                direct_records.append({**row, "category_code": category,
                                       "interaction_type": CATEGORIES[category],
                                       "source_sheet": sheet_name})
                seen_direct.add(key)

    # Shared official language is retained as its own interaction type.
    for row in sheet_rows(source["Language Tie Pairs"]):
        add_edge(
            edges, "C4", row.get("city_a"), row.get("city_b"), "L1",
            f"Shared official language(s): {row.get('shared_official_languages')}",
        )
        legacy_direct_pairs.add(tuple(sorted((row.get("city_a"), row.get("city_b")))))

    # Attach source URLs to regime-derived relations through membership records.
    regime_urls = defaultdict(set)
    for row in sheet_rows(source["Regime Memberships"]):
        if row.get("source_url"):
            regime_urls[row.get("historical_regime")].add(row.get("source_url"))

    legacy_regime_pairs = set()
    for row in sheet_rows(source["Shared Regime Pairs"]):
        legacy_regime_pairs.add(tuple(sorted((row.get("city_a"), row.get("city_b")))))
        regimes = [item.strip() for item in str(row.get("shared_regimes") or "").split(";") if item.strip()]
        for regime in regimes:
            targets = []
            if regime in COLONIAL_IMPERIAL_REGIMES:
                targets.append("C1")
            if regime in CIVILIZATIONAL_REGIMES:
                targets.append("C2")
            if regime in TRADE_REGIMES:
                targets.append("C5")
            for category in targets:
                urls = sorted(regime_urls.get(regime, [])) or [None]
                for url in urls:
                    add_edge(
                        edges, category, row.get("city_a"), row.get("city_b"),
                        f"Shared regime: {regime}", f"Both cities are recorded in {regime}.", url,
                    )

    category_pairs = {code: set() for code in CATEGORIES}
    for category, pair in edges:
        category_pairs[category].add(pair)
    all_pairs = set().union(*category_pairs.values())

    output = Workbook()
    output.remove(output.active)

    readme = output.create_sheet("README")
    readme_rows = [
        ["138-city cultural-homology interaction layers"],
        [],
        ["Item", "Definition"],
        ["City universe", "138 cities ordered as in the source City List."],
        ["Analytical structure", "Five non-exclusive interaction types; a city pair may have multiple labels."],
        ["C1", CATEGORIES["C1"]],
        ["C2", CATEGORIES["C2"]],
        ["C3", CATEGORIES["C3"]],
        ["C4", CATEGORIES["C4"]],
        ["C5", CATEGORIES["C5"]],
        ["Matrix coding", "1 indicates a documented or membership-derived interaction; 0 indicates no verified record and is not proof of absence."],
        ["Diagonal", "Set to 0 because self-pairs are not analyzed."],
        ["Direction", "Matrices are symmetric; directed source/target information is retained in Evidence Catalog."],
        ["Overlap", "Categories are multi-label, particularly for imperial trade networks and Ottoman ties."],
        ["Panel design", "The accompanying plot uses one panel with one CI distribution per interaction type."],
        ["Compatibility sheets", "The hidden Direct Tie Matrix and Shared Regime Matrix are retained only so legacy analysis scripts remain reproducible."],
        [],
        ["Quality-control item", "Value"],
        ["Cities", len(cities)],
        ["Any-type city pairs", len(all_pairs)],
        *[[f"{code} city pairs", len(category_pairs[code])] for code in CATEGORIES],
    ]
    for row in readme_rows:
        readme.append(row)
    readme.column_dimensions["A"].width = 28
    readme.column_dimensions["B"].width = 120
    readme["A1"].font = Font(size=16, bold=True, color="243447")
    readme.freeze_panes = "A3"

    city_ws = output.create_sheet("City List")
    copy_table(source["City List"], city_ws)
    style_table(city_ws)
    for width, col in zip((10, 16, 24, 24), "ABCD"):
        city_ws.column_dimensions[col].width = width

    pairs_ws = output.create_sheet("Interaction Pairs")
    pairs_ws.append([
        "pair_id", "city_a", "city_b", "category_code", "interaction_type",
        "evidence_basis", "evidence_detail", "confidence", "source_url",
    ])
    pair_id = 0
    for category in CATEGORIES:
        for pair in sorted(category_pairs[category]):
            pair_id += 1
            record = edges[(category, pair)]
            pairs_ws.append([
                pair_id, pair[0], pair[1], category, CATEGORIES[category],
                " | ".join(sorted(record["bases"])),
                " | ".join(sorted(record["details"])),
                "; ".join(sorted(record["confidences"])),
                " | ".join(sorted(record["urls"])),
            ])
    style_table(pairs_ws)
    widths = [10, 22, 22, 14, 44, 42, 90, 12, 70]
    for index, width in enumerate(widths, 1):
        pairs_ws.column_dimensions[get_column_letter(index)].width = width

    add_matrix_sheet(output, "Any Interaction Matrix", cities, all_pairs)
    for code, sheet_name in MATRIX_SHEETS.items():
        add_matrix_sheet(output, sheet_name, cities, category_pairs[code])

    # Hidden compatibility layers prevent the pre-existing validation scripts
    # from breaking, while the visible supplementary structure remains five-type.
    add_matrix_sheet(output, "Direct Tie Matrix", cities, legacy_direct_pairs, diagonal=1)
    output["Direct Tie Matrix"].sheet_state = "hidden"
    add_matrix_sheet(output, "Shared Regime Matrix", cities, legacy_regime_pairs, diagonal=1)
    output["Shared Regime Matrix"].sheet_state = "hidden"

    evidence_ws = output.create_sheet("Evidence Catalog")
    evidence_headers = [
        "source_sheet", "source_city", "target_city", "original_tie_code",
        "category_code", "interaction_type", "evidence_summary", "confidence", "source_url",
    ]
    evidence_ws.append(evidence_headers)
    for row in direct_records:
        evidence_ws.append([
            row.get("source_sheet"), row.get("source_city"), row.get("target_city"),
            row.get("tie_code"), row.get("category_code"), row.get("interaction_type"),
            row.get("evidence_summary"), row.get("confidence"), row.get("source_url"),
        ])
    style_table(evidence_ws)
    for index, width in enumerate((26, 22, 22, 18, 14, 44, 100, 12, 70), 1):
        evidence_ws.column_dimensions[get_column_letter(index)].width = width

    rules_ws = output.create_sheet("Classification Rules")
    rules_ws.append(["input_record", "assigned_category", "rationale"])
    for code, category in DIRECT_CODE_TO_CATEGORY.items():
        rules_ws.append([code, f"{category}: {CATEGORIES[category]}", "Mapped from the original evidence codebook."])
    for regime in sorted(COLONIAL_IMPERIAL_REGIMES):
        rules_ws.append([f"Shared regime: {regime}", f"C1: {CATEGORIES['C1']}", "Colonial or imperial membership."])
    for regime in sorted(CIVILIZATIONAL_REGIMES):
        rules_ws.append([f"Shared regime: {regime}", f"C2: {CATEGORIES['C2']}", "Shared civilizational or ideological sphere."])
    for regime in sorted(TRADE_REGIMES):
        rules_ws.append([f"Shared regime: {regime}", f"C5: {CATEGORIES['C5']}", "Chartered trading-company network."])
    rules_ws.append(["L1", f"C4: {CATEGORIES['C4']}", "At least one shared official or de facto official language."])
    style_table(rules_ws)
    for index, width in enumerate((48, 52, 62), 1):
        rules_ws.column_dimensions[get_column_letter(index)].width = width

    # Retain the source membership tables required to audit or rebuild edges.
    for source_name in ("Regime Memberships", "Shared Regime Pairs", "Official Languages", "Language Tie Pairs"):
        ws = output.create_sheet(source_name)
        copy_table(source[source_name], ws)
        style_table(ws)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = min(
                70, max(14, max(len(str(ws.cell(row, col).value or "")) for row in range(1, min(ws.max_row, 250) + 1)) + 2)
            )

    # Save atomically so interruption cannot corrupt the only local workbook.
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=WORKBOOK.parent) as tmp:
        temporary_path = Path(tmp.name)
    try:
        output.save(temporary_path)
        check = load_workbook(temporary_path, read_only=True, data_only=True)
        for title in ["Interaction Pairs", "Any Interaction Matrix", *MATRIX_SHEETS.values()]:
            if title not in check.sheetnames:
                raise RuntimeError(f"Output validation failed: missing {title}")
        check.close()
        shutil.move(str(temporary_path), WORKBOOK)
        os.chmod(WORKBOOK, 0o644)
    finally:
        temporary_path.unlink(missing_ok=True)

    print(f"Updated {WORKBOOK}")
    for code, label in CATEGORIES.items():
        print(f"{code}: {len(category_pairs[code]):4d} pairs  {label}")
    print(f"Any type: {len(all_pairs):4d} unique pairs")


if __name__ == "__main__":
    main()
